from openpyxl import load_workbook
# https://openpyxl.readthedocs.io/en/stable/
# https://letscode-academy.com/blog/aprenda-a-integrar-python-e-excel/

estados = [('ac', 'Acre.xlsx'),
           ('al', 'Alagoas.xlsx'),
           ('ap', 'Amapá.xlsx'),
           ('am', 'Amazonas.xlsx'),
           ('ba', 'Bahia.xlsx'),
           ('ce', 'Ceará.xlsx'),
           ('df', 'Distrito Federal.xlsx'),
           ('es', 'Espírito Santo.xlsx'),
           ('go', 'Goiás.xlsx'),
           ('ma', 'Maranhão.xlsx'),
           ('mt', 'Mato Grosso.xlsx'),
           ('ms', 'Mato Grosso do Sul.xlsx'),
           ('mg', 'Minas Gerais.xlsx'),
           ('pa', 'Pará.xlsx'),
           ('pb', 'Paraíba.xlsx'),
           ('pr', 'Paraná.xlsx'),
           ('pe', 'Pernambuco.xlsx'),
           ('pi', 'Piauí.xlsx'),
           ('rj', 'Rio de Janeiro.xlsx'),
           ('rn', 'Rio Grande do Norte.xlsx'),
           ('rs', 'Rio Grande do Sul.xlsx'),
           ('ro', 'Rondônia.xlsx'),
           ('rr', 'Roraima.xlsx'),
           ('sc', 'Santa Catarina.xlsx'),
           ('sp', 'São Paulo.xlsx'),
           ('se', 'Sergipe.xlsx'),
           ('to', 'Tocantins.xlsx')]

areas = ['Engenharia, produção e construção', 'Agricultura e veterinária', 'Ciências, matemática e computação', 'Saúde e bem estar social', 'Serviços']

main_file = 'Consolidado.xlsx'
wb_main = load_workbook(filename=main_file)
ws_main = wb_main.active
wb_main_row = 2

lista_cod_curso = []

for e in estados:
    wb_bd = load_workbook(filename=e[1])
    print(e[1])
# iterar sobre todas as rows presentes na worksheet
# deletar as colunas que não precisamos?
# da pra iterar sobre tudo, mas o custo é um pouco maior.... ficando dentro de dezenas, tá tranquilo

    ws = wb_bd.active
    for row in ws.iter_rows(min_row=7, min_col=21, max_col=21, max_row=ws.max_row):
        for cell in row:
            for a in areas:
                if cell.value == a:
                    # se o par instituição de ensino/curso não está na lista de IES/curso
                    # então copia o registro pra consolidado
                    # e adiciona a IES e curso para a lista.
                    if ws.cell(column = 4, row = cell.row).value not in lista_cod_curso:
                        ws_main.cell(column = 1, row = wb_main_row).value = ws.cell(column = 5, row = cell.row).value
                        ws_main.cell(column = 2, row = wb_main_row).value = ws.cell(column = 6, row = cell.row).value
                        ws_main.cell(column = 3, row = wb_main_row).value = ws.cell(column = 7, row = cell.row).value
                        ws_main.cell(column = 4, row = wb_main_row).value = ws.cell(column = 16, row = cell.row).value
                        ws_main.cell(column = 5, row = wb_main_row).value = ws.cell(column = 21, row = cell.row).value
                        ws_main.cell(column = 6, row = wb_main_row).value = ws.cell(column = 4, row = cell.row).value
                        ws_main.cell(column = 7, row = wb_main_row).value = e[0]
                        lista_cod_curso.append(ws.cell(column = 4, row = cell.row).value)
                        wb_main_row+=1
                    else:
                        # ws_main.cell(column = 7, row = wb_main_row).value = str(ws_main.cell(column = 7, row = wb_main_row).value) + e[0]
                        ws_main.cell(column = 8, row = wb_main_row).value = 'duplicado!!'
                        # print(ws_main.cell(column = 6, row = wb_main_row).value)
                    # se o par JÁ está na lista
                    # então adiciona o estado a coluna de estados
                    # e não avança a wb_main_row

            # if cell.col_idx == 5:
            #     ws_main.cell(column = 1, row = wb_main_row).value = cell.value
            #     # print(cell.value)
            #     # print(ws_main.cell(column = 1, row = wb_main_row).value)
            #     # copiar os conteudos das celulas desejadas pro destino no main file
            #     # copiar direto pra um 'main' worksheet ou separar em worksheets? se for o segundo, teria que definir um critério... switch case?
            # if cell.col_idx == 6:
            #     ws_main.cell(column = 2, row = wb_main_row).value = cell.value
            #     # print(cell.value)
            # if cell.col_idx == 7:
            #     ws_main.cell(column = 3, row = wb_main_row).value = cell.value
            #     # print(cell.value)
            # if cell.col_idx == 16:
            #     ws_main.cell(column = 4, row = wb_main_row).value = cell.value
            #     # print(cell.value)
            # if cell.col_idx == 21:
            #     ws_main.cell(column = 5, row = wb_main_row).value = cell.value
            #     # print(cell.value)
            # # cell.value

            # ws_main.cell(column = 6, row = wb_main_row).value = e[0]
            # wb_main_row+=1
wb_main.save('Consolidado.xlsx')